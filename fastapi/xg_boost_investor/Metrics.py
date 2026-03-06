import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import warnings

warnings.filterwarnings('ignore')


def analyze_portfolio(portfolio_data):
    # Convert to DataFrame
    df = pd.DataFrame(portfolio_data, columns=['Date', 'PortfolioValue'])
    df['Date'] = pd.to_datetime(df['Date'])
    df = df.sort_values('Date').reset_index(drop=True)

    # Calculate daily returns
    df['DailyReturn'] = df['PortfolioValue'].pct_change()
    df['ROI'] = (df['PortfolioValue'] / df['PortfolioValue'].iloc[0] - 1) * 100

    # Calculate cumulative maximum for drawdown
    df['CumulativeMax'] = df['PortfolioValue'].cummax()
    df['Drawdown'] = (df['PortfolioValue'] - df['CumulativeMax']) / df['CumulativeMax'] * 100

    # Metrics calculation
    total_return = (df['PortfolioValue'].iloc[-1] / df['PortfolioValue'].iloc[0] - 1) * 100
    annual_return = (total_return / 100) ** (365 / len(df)) - 1

    daily_returns = df['DailyReturn'].dropna()
    volatility = daily_returns.std() * np.sqrt(252)  # Annualized

    risk_free_rate = 0.04  # 4% annual risk-free rate
    excess_return = annual_return - risk_free_rate
    sharpe_ratio = excess_return / volatility if volatility != 0 else 0

    # Sortino Ratio (only downside volatility)
    downside_returns = daily_returns[daily_returns < 0]
    downside_volatility = downside_returns.std() * np.sqrt(252) if len(downside_returns) > 0 else 0
    sortino_ratio = excess_return / downside_volatility if downside_volatility != 0 else 0

    # Maximum Drawdown
    max_drawdown = df['Drawdown'].min()

    # Calmar Ratio
    calmar_ratio = annual_return / abs(max_drawdown) * 100 if max_drawdown != 0 else 0

    # Win Rate
    positive_days = len(daily_returns[daily_returns > 0])
    total_days = len(daily_returns)
    win_rate = (positive_days / total_days * 100) if total_days > 0 else 0

    # Profit Factor
    gains = daily_returns[daily_returns > 0].sum()
    losses = abs(daily_returns[daily_returns < 0].sum())
    profit_factor = gains / losses if losses != 0 else 0

    # Recovery Factor
    total_days_traded = len(df) - 1
    total_profit = df['PortfolioValue'].iloc[-1] - df['PortfolioValue'].iloc[0]
    recovery_factor = total_profit / abs(max_drawdown * df['CumulativeMax'].max() / 100) if max_drawdown != 0 else 0

    metrics = {
        'Total Return (%)': total_return,
        'Annual Return (%)': annual_return * 100,
        'Volatility (%)': volatility * 100,
        'Sharpe Ratio': sharpe_ratio,
        'Sortino Ratio': sortino_ratio,
        'Maximum Drawdown (%)': max_drawdown,
        'Calmar Ratio': calmar_ratio,
        'Win Rate (%)': win_rate,
        'Profit Factor': profit_factor,
        'Recovery Factor': recovery_factor,
    }

    return df, metrics


def create_visualizations(df, title, year, output_dir='./'):
    # Chart 1: Portfolio Value Over Time
    fig, ax = plt.subplots(figsize=(12, 6))
    ax.plot(df['Date'], df['PortfolioValue'], linewidth=2, color='#1f77b4')
    ax.fill_between(df['Date'], df['PortfolioValue'], alpha=0.3, color='#1f77b4')
    ax.set_xlabel('Date', fontsize=12, fontweight='bold')
    ax.set_ylabel('Portfolio Value ($)', fontsize=12, fontweight='bold')
    ax.set_title('Portfolio Value Over Time', fontsize=14, fontweight='bold')
    ax.grid(True, alpha=0.3)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'${x:,.0f}'))
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.savefig(f'{output_dir}portfolio_value_{title}_{year}.png', dpi=300, bbox_inches='tight')
    plt.close()

    # Chart 2: ROI Over Time
    fig, ax = plt.subplots(figsize=(12, 6))
    ax.plot(df['Date'], df['ROI'], linewidth=2, color='#2ca02c')
    ax.fill_between(df['Date'], df['ROI'], alpha=0.3, color='#2ca02c')
    ax.axhline(y=0, color='red', linestyle='--', alpha=0.5)
    ax.set_xlabel('Date', fontsize=12, fontweight='bold')
    ax.set_ylabel('ROI (%)', fontsize=12, fontweight='bold')
    ax.set_title('Return on Investment (ROI) Over Time', fontsize=14, fontweight='bold')
    ax.grid(True, alpha=0.3)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:.1f}%'))
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.savefig(f'{output_dir}roi_over_time_{title}_{year}.png', dpi=300, bbox_inches='tight')
    plt.close()


def export_metrics(metrics, title, year, output_dir='./'):
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Portfolio Metrics'

    # Title
    sheet['A1'] = 'Portfolio Performance Metrics'
    sheet['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    sheet['A1'].fill = PatternFill(start_color='1f77b4', end_color='1f77b4', fill_type='solid')
    sheet.merge_cells('A1:B1')
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Headers
    sheet['A3'] = 'Metric'
    sheet['B3'] = 'Value'
    for cell in ['A3', 'B3']:
        sheet[cell].font = Font(bold=True, color='FFFFFF')
        sheet[cell].fill = PatternFill(start_color='4472c4', end_color='4472c4', fill_type='solid')
        sheet[cell].alignment = Alignment(horizontal='center', vertical='center')

    # Add metrics
    row = 4
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for metric_name, value in metrics.items():
        sheet[f'A{row}'] = metric_name
        sheet[f'B{row}'] = value
        sheet[f'B{row}'].number_format = '0.00'

        # Alternating row colors
        if row % 2 == 0:
            fill_color = PatternFill(start_color='e7f0f7', end_color='e7f0f7', fill_type='solid')
            sheet[f'A{row}'].fill = fill_color
            sheet[f'B{row}'].fill = fill_color

        # Apply border
        sheet[f'A{row}'].border = border
        sheet[f'B{row}'].border = border
        sheet[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        sheet[f'B{row}'].alignment = Alignment(horizontal='right', vertical='center')

        row += 1

    # Column widths
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 18

    # Add description section
    row += 2
    sheet[f'A{row}'] = 'Metric Descriptions'
    sheet[f'A{row}'].font = Font(bold=True, size=12)

    descriptions = {
        'Total Return': 'Percentage gain/loss from start to end of period',
        'Annual Return': 'Annualized return over the period',
        'Volatility': 'Annualized standard deviation of daily returns',
        'Sharpe Ratio': 'Risk-adjusted return (higher is better, >1 is good)',
        'Sortino Ratio': 'Return per unit of downside risk (higher is better)',
        'Max Drawdown': 'Largest peak-to-trough decline (more negative is worse)',
        'Calmar Ratio': 'Annual return divided by max drawdown (higher is better)',
        'Win Rate': 'Percentage of days with positive returns',
        'Profit Factor': 'Ratio of gross profit to gross loss (>1 is profitable)',
        'Recovery Factor': 'Net profit divided by max drawdown (higher is better)',
    }

    row += 1
    for desc_name, desc_text in descriptions.items():
        sheet[f'A{row}'] = f'{desc_name}:'
        sheet[f'A{row}'].font = Font(bold=True, size=10)
        sheet[f'B{row}'] = desc_text
        sheet[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        sheet.row_dimensions[row].height = 30
        row += 1

    wb.save(f'{output_dir}portfolio_metrics_{title}_{year}.xlsx')


# Example usage with sample data
if __name__ == '__main__':
    # Generate sample portfolio data (1 year)
    dates = pd.date_range(start='2023-01-01', end='2023-12-31', freq='D')

    # Simulate portfolio values with some trend and volatility
    np.random.seed(42)
    returns = np.random.normal(0.0005, 0.015, len(dates))
    portfolio_values = 100000 * np.exp(np.cumsum(returns))

    # Create tuples
    portfolio_data = list(zip(dates, portfolio_values))

    # Analyze
    df, metrics = analyze_portfolio(portfolio_data)

    # Export
    create_visualizations(df, "Title", 2023)
    export_metrics(metrics, "Title", 2023)

    print("Analysis complete!")
    print(f"\nPortfolio Metrics:")
    for metric, value in metrics.items():
        print(f"  {metric}: {value:.2f}")